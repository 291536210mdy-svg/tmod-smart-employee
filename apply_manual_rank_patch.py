import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


RANK_PREFIX_RE = re.compile(r"^本奖项排名第\d+位。")


def normalize_reason_rank(reason, rank):
    text = str(reason or "").strip()
    text = RANK_PREFIX_RE.sub("", text)
    return f"本奖项排名第{rank}位。{text}" if text else f"本奖项排名第{rank}位。"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--report", required=True)
    args = parser.parse_args()

    source_path = Path(args.source)
    output_path = Path(args.output)
    report_path = Path(args.report)

    wb = load_workbook(source_path)
    ws = wb.worksheets[0]
    wd = wb.worksheets[1]

    headers = {wd.cell(1, col).value: col for col in range(1, wd.max_column + 1)}
    detail_by_cid = {}
    for row in range(2, wd.max_row + 1):
        cid = wd.cell(row, headers["candidate_id"]).value
        if cid:
            detail_by_cid[cid] = {
                "row": row,
                "rank": wd.cell(row, headers["award_rank"]).value,
                "status": wd.cell(row, headers["recommendation_status"]).value,
                "reason": wd.cell(row, headers["ranking_reason"]).value or "",
            }

    current_global_by_rank = {
        1: "row_00018",
        2: "row_00004",
        3: "row_00019",
        4: "row_00005",
        5: "row_00039",
        6: "row_00020",
        7: "row_00022",
        8: "row_00037",
        9: "row_00021",
    }
    current_corp_by_rank = {
        1: "row_00036",
        2: "row_00035",
        3: "row_00012",
    }
    result_values_by_cid = {}

    def collect_values(marker, rank_to_cid):
        current = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value:
                current = ws.cell(row, 1).value
            if current and marker in current:
                rank = ws.cell(row, 2).value
                cid = rank_to_cid.get(int(rank)) if isinstance(rank, int) or str(rank).isdigit() else None
                if cid:
                    result_values_by_cid[cid] = [ws.cell(row, col).value for col in range(2, 10)]

    collect_values("Global Business Breakthrough", current_global_by_rank)
    collect_values("Corporate Transformational Growth", current_corp_by_rank)

    manual_reasons = {
        "row_00037": (
            "本奖项排名第1位。该项目由复星医药创新药事业部与复星战略发展部联合推动，"
            "历时2年与阿联酋主权基金ADQ及Arcera达成战略合作，打开中东主权资本与生命科学产业协同通道。"
            "合作覆盖创新药早期研发融资、成熟药海外授权License out、罕见病药物当地JV等关键方向，"
            "兼具全球化市场突破、资本合作突破和后续管线商业化平台价值，战略牵引性与集团级示范意义突出。"
        ),
        "row_00018": (
            "本奖项排名第2位。综合考虑复宏汉霖内部战略优先级及项目里程碑，"
            "HLX10围术期胃癌中国上市申请实现6个月内获批、审评期间零发补，显著优于肿瘤创新药通常审评周期，"
            "并连续完成突破性治疗认定、优先审评纳入及上市获批。虽全球化及本土以外业务证据仍需补充，"
            "但注册效率和产品商业化推进价值突出。结合复宏汉霖内部战略优先级。"
        ),
        "row_00035": (
            "本奖项排名第1位。综合复星健康内部战略优先级及正常评审证据，"
            "上海星晨儿童医院在收入、减亏和现金流改善上均有量化表现：2026年1-5月累计收入2,937万元，"
            "同比增长36.87%，净利润同比减亏583万元，现金流净支出同比减少953万元；"
            "同时精神科病房、儿童心理健康项目、生长发育门诊及24小时急诊等业务拓展体现增长质量。"
            "结合复星健康内部战略优先级。"
        ),
        "row_00036": (
            "本奖项排名第2位。宿迁钟吾医院正常评审分较高，收入、利润和现金流指标表现扎实，"
            "总收入1.42亿元、净利润同比增长273%、自由现金流同比增长145%；"
            "但复星健康内部本奖项优先推荐上海星晨儿童医院，宿迁作为同奖项复星健康候选在原有槽位内顺延至第2位，"
            "整体排序兼顾正常评审表现与复星健康内部战略优先级。结合复星健康内部战略优先级。"
        ),
    }

    global_order = [
        "row_00037",
        "row_00018",
        "row_00004",
        "row_00019",
        "row_00005",
        "row_00039",
        "row_00020",
        "row_00022",
        "row_00021",
    ]
    corp_order = ["row_00035", "row_00036", "row_00012"]

    def award_rows(marker):
        rows = []
        current = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value:
                current = ws.cell(row, 1).value
            if current and marker in current:
                rows.append(row)
        return rows

    changed = []

    def write_order(marker, order):
        rows = award_rows(marker)
        if len(rows) != len(order):
            raise RuntimeError(f"{marker} row count mismatch: sheet={len(rows)} order={len(order)} rows={rows}")
        for rank, (row, cid) in enumerate(zip(rows, order), start=1):
            values = list(result_values_by_cid[cid])
            values[0] = rank
            values[7] = manual_reasons.get(cid, normalize_reason_rank(detail_by_cid[cid]["reason"], rank))
            for col, value in enumerate(values, start=2):
                ws.cell(row, col).value = value
            changed.append({
                "candidate_id": cid,
                "marker": marker,
                "result_row": row,
                "new_rank": rank,
                "subject": values[1],
                "new_reason": values[7],
            })

    write_order("Global Business Breakthrough", global_order)
    write_order("Corporate Transformational Growth", corp_order)

    status_overrides = {
        "row_00037": "recommended",
        "row_00018": "recommended",
        "row_00004": "not_selected",
        "row_00019": "not_selected",
        "row_00005": "not_selected",
        "row_00039": "not_selected",
        "row_00020": "needs_review",
        "row_00022": "not_selected",
        "row_00021": "not_selected",
        "row_00035": "recommended",
        "row_00036": "needs_review",
        "row_00012": "recommended",
    }
    new_ranks = {cid: i for i, cid in enumerate(global_order, start=1)}
    new_ranks.update({cid: i for i, cid in enumerate(corp_order, start=1)})
    for cid, rank in new_ranks.items():
        row = detail_by_cid[cid]["row"]
        reason = manual_reasons.get(cid, normalize_reason_rank(detail_by_cid[cid]["reason"], rank))
        wd.cell(row, headers["award_rank"]).value = rank
        wd.cell(row, headers["recommendation_status"]).value = status_overrides.get(cid, detail_by_cid[cid]["status"])
        wd.cell(row, headers["ranking_reason"]).value = reason
        if cid in {"row_00037", "row_00035", "row_00036"}:
            wd.cell(row, headers["ranking_reason_source"]).value = "manual_patch"
            wd.cell(row, headers["ranking_reason_json"]).value = json.dumps(
                {"manual_patch": True, "reason": "user requested strategic rank adjustment"},
                ensure_ascii=False,
            )

    wb.save(output_path)
    report = {
        "source_xlsx": str(source_path),
        "output_xlsx": str(output_path),
        "global_order": global_order,
        "corporate_order": corp_order,
        "changed": changed,
    }
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"saved": str(output_path), "report": str(report_path), "changed_rows": len(changed)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
