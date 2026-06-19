import argparse
import copy
import datetime as dt
import difflib
import json
import os
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Protocol

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


WORKSPACE = Path(__file__).resolve().parent
DEFAULT_ENV_FILES = [
    WORKSPACE / ".env",
    Path(r"E:\Agent学习\my_project\.env"),
]
DEFAULT_INPUT = WORKSPACE / "sample_input.xlsx"
DEFAULT_TEMPLATE = WORKSPACE / "评选结果输出格式.xlsx"
DEFAULT_OUTPUT_DIR = WORKSPACE / "outputs"
DEFAULT_AWARD_CONFIG = WORKSPACE / "award_config.json"
DEFAULT_LEADERSHIP_PRIORITIES = []
TARGET_HEADERS = ["奖项名称", "序号", "主体", "所属BU", "所属PL", "提报人", "团队负责人", "事迹", "排名理由"]
INPUT_REQUIRED_HEADERS = ["申报批次", "申报项目", "所属BU", "申报主体", "姓名", "推荐人姓名", "筛选人/对接人", "申报理由"]
INPUT_HEADER_ALIASES = {
    "申报项目": ["奖项名称"],
    "申报主体": ["主体"],
    "申报理由": ["事迹"],
    "姓名": ["提报人"],
    "推荐人姓名": ["提报人"],
}
PLACEHOLDER_VALUES = {
    "待补充", "待填写", "请填写", "无", "暂无", "N/A", "NA", "-",
    "此处有一段话", "这里有一段话", "high", "medium", "strong", "weak", "missing",
}
ENABLE_LEADERSHIP_PRIORITY = True
LEADERSHIP_PRIORITY_WEIGHT = 0.8
NORMAL_REVIEW_WEIGHT = 0.2
FIXED_LEADERSHIP_PRIORITY_WEIGHT = 1.0
FIXED_NORMAL_REVIEW_WEIGHT = 0.0
FIXED_LEADERSHIP_PRIORITY_RULES = [
    {
        "source_label": "复宏汉霖",
        "award_name": "全球业务突破奖 Global Business Breakthrough Award",
        "subjects": [
            "复宏汉霖 -药政注册部",
            {"subject": "复宏汉霖  全球产品开发部", "aliases": ["复宏汉霖  HLX10项目组"]},
            {"subject": "复宏汉霖  HLX11 产品组", "aliases": ["复宏汉霖  HLX11项目组"]},
            "复宏汉霖商务拓展部",
            {"subject": "复宏汉霖 税务团队", "aliases": ["复宏汉霖 财务部"]},
        ],
    },
    {
        "source_label": "复宏汉霖",
        "award_name": "AI价值领航奖 AI Value Navigator Award",
        "subjects": [
            {"subject": "复宏汉霖 全球创新中心", "aliases": ["复宏汉霖 HAI Club"]},
            "复宏汉霖 数智创新部",
        ],
    },
    {
        "source_label": "复星健康",
        "award_name": "企业经营乘长奖 Corporate Transformational Growth Award",
        "subjects": [
            "上海星晨儿童医院",
            "宿迁钟吾医院院委会",
        ],
    },
    {
        "source_label": "复星健康",
        "award_name": "AI价值领航奖 AI Value Navigator Award",
        "subjects": [
            "广州复星禅诚医院",
            "中国药科大学附属重庆星荣整形外科医院新媒体营销部",
        ],
    },
]
RANK_ONE_BAD_REASON_MARKERS = (
    "与的候选人相比",
    "与候选人相比",
    "与排名第",
    "相较于排名第",
    "相比排名第",
    "与排名更高",
    "与排名更前",
    "与排名前列",
    "与排名前两位",
    "与排名第一",
    "与排名靠前",
    "相较于排名更高",
    "相较于排名前列",
    "相较于排名靠前",
    "排名更高的候选人",
    "排名靠前的候选人",
    "排名靠前的团队",
    "更高的候选人",
    "靠前的候选人",
    "靠前的团队",
    "更高排名",
    "排名居中",
    "排名中等",
    "当前排名",
    "略逊",
    "亮点不足",
    "亮点不够突出",
)
GLOBAL_BAD_REASON_MARKERS = ("与的候选人相比", "与候选人相比")
GRADE_POINTS = {
    "strong": 100,
    "medium": 70,
    "weak": 35,
    "missing": 0,
}
DEFAULT_AWARD_CONFIG_DATA = {
    "default": {
        "quota": 2,
        "weights": {
            "rule_match": 0.25,
            "quantitative": 0.20,
            "value_impact": 0.25,
            "innovation": 0.15,
            "strategy_align": 0.15,
        },
        "risk_penalty_per_flag": 5,
        "max_risk_penalty": 15,
        "tie_break_keywords": ["一线", "海外"],
    },
    "awards": {},
}


@dataclass
class ReviewBatchConfig:
    input_path: Path
    output_dir: Path
    template_path: Path = DEFAULT_TEMPLATE
    award_config_path: Path = DEFAULT_AWARD_CONFIG
    leadership_priority_paths: list[Path] = field(default_factory=list)
    enable_leadership_priority: bool = True
    top_n: int = 2
    award_filters: list[str] = field(default_factory=list)
    limit: int = 0
    sleep: float = 0.2
    timeout: int = 120
    dry_run: bool = False
    dify_base_url: str = ""
    dify_review_api_key: str = ""
    dify_ranking_reason_api_key: str = ""
    dify_user: str = "review-batch"


@dataclass
class ReviewBatchResult:
    output_dir: Path
    xlsx_path: Path
    raw_jsonl_path: Path
    internal_pack_path: Path
    completion_path: Path
    qa_report_path: Path
    qa_passed: bool
    expected_rows: int
    processed_rows: int
    award_counts: dict
    qa_report: dict


class EventSink(Protocol):
    def emit(
        self,
        event_type: str,
        *,
        message: str = "",
        level: str = "info",
        progress: tuple[int, int] | None = None,
        payload: dict | None = None,
    ) -> None:
        ...


class NullEventSink:
    def emit(self, *args, **kwargs):
        return None


class RunCancelled(Exception):
    pass


def load_env_files(paths):
    for path in paths:
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8-sig").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            os.environ.setdefault(key, value)


def merge_dict(base, override):
    merged = copy.deepcopy(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = merge_dict(merged[key], value)
        else:
            merged[key] = value
    return merged


def load_award_config(path, top_n):
    config = copy.deepcopy(DEFAULT_AWARD_CONFIG_DATA)
    if path.exists():
        loaded = json.loads(path.read_text(encoding="utf-8-sig"))
        config = merge_dict(config, loaded)
    if top_n:
        config.setdefault("default", {})["quota"] = top_n
    return config


def get_award_config(config, award_name):
    award_config = copy.deepcopy(config.get("default", {}))
    for keyword, override in config.get("awards", {}).items():
        if keyword and keyword in award_name:
            award_config = merge_dict(award_config, override)
    return award_config


def cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def find_header_row(ws):
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        values = [cell_to_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        has_old_headers = "申报项目" in values and "申报理由" in values
        has_new_headers = "奖项名称" in values and "事迹" in values
        if has_old_headers or has_new_headers:
            return row_idx, values
    raise ValueError("没有找到包含“申报项目/申报理由”或“奖项名称/事迹”的表头行")


def normalize_input_values(values):
    normalized = dict(values)
    for canonical, aliases in INPUT_HEADER_ALIASES.items():
        if normalized.get(canonical):
            continue
        for alias in aliases:
            if normalized.get(alias):
                normalized[canonical] = normalized[alias]
                break
    return normalized


def read_excel_records(input_path):
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active
    header_row, headers = find_header_row(ws)
    header_positions = {name: idx + 1 for idx, name in enumerate(headers) if name}
    missing = [
        name
        for name in ["申报项目", "申报理由"]
        if name not in header_positions and not any(alias in header_positions for alias in INPUT_HEADER_ALIASES.get(name, []))
    ]
    if missing:
        raise ValueError(f"输入表缺少必要列：{', '.join(missing)}")

    records = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        values = {
            header: cell_to_text(ws.cell(row_idx, col_idx).value)
            for header, col_idx in header_positions.items()
        }
        if not any(values.values()):
            continue
        values = normalize_input_values(values)
        records.append({"excel_row": row_idx, "values": values})
    return records


def infer_award_type(row):
    award_name = row.get("申报项目", "")
    subject = row.get("申报主体", "")
    name = row.get("姓名", "")
    team_words = ["团队", "小组", "集体", "项目组", "Team", "team"]
    if any(word in award_name or word in subject for word in team_words):
        return "团队奖"
    if name:
        return "个人奖"
    return "未知"


def mask_reason(reason, row):
    masked = reason or ""
    replacements = {
        "姓名": "[候选人]",
        "推荐人姓名": "[推荐人]",
        "筛选人/对接人": "[对接人]",
        "申报主体": "[申报主体]",
        "公司": "[公司]",
        "所属BU": "[所属BU]",
        "所属BG/AMC": "[所属BG]",
    }
    pairs = []
    for field, token in replacements.items():
        value = row.get(field, "")
        if len(value) >= 2:
            pairs.append((value, token))
    for value, token in sorted(pairs, key=lambda item: len(item[0]), reverse=True):
        masked = masked.replace(value, token)
    return masked


def build_workflow_inputs(record):
    row = record["values"]
    reason = row.get("申报理由", "")
    masked_reason = mask_reason(reason, row)
    if not masked_reason.strip() or masked_reason.strip() in PLACEHOLDER_VALUES:
        review_reason = f"候选人申报理由（唯一证据来源）：无具体申报理由。原始内容为占位或空泛文本：{masked_reason}"
    else:
        review_reason = (
            "候选人申报理由（唯一可作为候选人证据的文本；"
            "不得把评优规则、评选标准或规则中的示例当作候选人证据）："
            f"{masked_reason}"
        )
    candidate_id = f"row_{record['excel_row']:05d}"
    return {
        "candidate_id": candidate_id,
        "batch_id": row.get("申报批次", ""),
        "award_name": row.get("申报项目", ""),
        "award_type": infer_award_type(row),
        "submission_reason_masked": review_reason,
        "submission_reason_full": f"完整申报理由：{reason}" if reason else "完整申报理由：",
        "raw_row_json": json.dumps(row, ensure_ascii=False),
    }


def workflow_run_endpoint(base_url):
    base = base_url.rstrip("/")
    if base.endswith("/v1"):
        return base + "/workflows/run"
    return base + "/v1/workflows/run"


def response_text_for_error(response, limit=2000):
    try:
        text = json.dumps(response.json(), ensure_ascii=False)
    except ValueError:
        text = response.text or ""
    text = text.strip()
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "..."


def input_lengths(inputs):
    return {key: len(str(value or "")) for key, value in inputs.items()}


def call_dify_workflow(base_url, api_key, inputs, user, timeout):
    endpoint = workflow_run_endpoint(base_url)
    payload = {
        "inputs": inputs,
        "response_mode": "blocking",
        "user": user,
    }
    response = requests.post(
        endpoint,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=timeout,
    )
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        error_detail = response_text_for_error(response)
        lengths = json.dumps(input_lengths(inputs), ensure_ascii=False)
        raise RuntimeError(
            f"Dify workflow HTTP {response.status_code}: {error_detail}; input_lengths={lengths}"
        ) from exc
    body = response.json()
    data = body.get("data", body)
    outputs = data.get("outputs", {})
    return body, outputs


def parse_json_maybe(text):
    if isinstance(text, dict):
        return text
    if not text:
        return {}
    raw = str(text).strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        candidates = []
        for start, char in enumerate(raw):
            if char != "{":
                continue
            depth = 0
            in_string = False
            escaped = False
            for end in range(start, len(raw)):
                current = raw[end]
                if escaped:
                    escaped = False
                    continue
                if current == "\\":
                    escaped = True
                    continue
                if current == '"':
                    in_string = not in_string
                    continue
                if in_string:
                    continue
                if current == "{":
                    depth += 1
                elif current == "}":
                    depth -= 1
                    if depth == 0:
                        candidates.append(raw[start : end + 1])
                        break
        parsed_candidates = []
        for candidate in reversed(candidates):
            try:
                parsed = json.loads(candidate)
            except json.JSONDecodeError:
                continue
            if isinstance(parsed, dict):
                parsed_candidates.append(parsed)
                if parsed.get("schema_version"):
                    return parsed
        if parsed_candidates:
            for parsed in parsed_candidates:
                if "evidence_grades" in parsed or "field_confidence" in parsed:
                    return parsed
            return parsed_candidates[0]
    return {}


def first_non_empty(*values, default=""):
    for value in values:
        if value and str(value).strip() not in PLACEHOLDER_VALUES:
            return value
    return default


def choose_field(candidates, default="待补充"):
    for value, source, confidence in candidates:
        if value and str(value).strip() not in PLACEHOLDER_VALUES:
            return value, {"source": source, "confidence": confidence or "", "value": value}
    return default, {"source": "fallback", "confidence": "low", "value": default}


def build_result_row(index, source_row, final_fields):
    row = source_row["values"]
    award_type = infer_award_type(row)
    confidence = final_fields.get("field_confidence", {}) if isinstance(final_fields.get("field_confidence"), dict) else {}
    team_leader_fallback = row.get("姓名", "") if award_type == "个人奖" else final_fields.get("team_leader", "")
    subject, subject_source = choose_field([
        (row.get("申报主体"), "excel.申报主体", "high"),
        (row.get("姓名"), "excel.姓名", "high"),
        (final_fields.get("subject"), "dify.final_fields.subject", confidence.get("subject")),
    ])
    submitter, submitter_source = choose_field([
        (row.get("姓名"), "excel.姓名", "high"),
        (row.get("推荐人姓名"), "excel.推荐人姓名", "high"),
        (row.get("筛选人/对接人"), "excel.筛选人/对接人", "medium"),
        (final_fields.get("submitter"), "dify.final_fields.submitter", confidence.get("submitter")),
    ])
    team_leader, team_leader_source = choose_field([
        (team_leader_fallback, "excel.姓名" if award_type == "个人奖" else "dify.final_fields.team_leader", "high" if award_type == "个人奖" else confidence.get("team_leader")),
        (final_fields.get("team_leader"), "dify.final_fields.team_leader", confidence.get("team_leader")),
    ])
    achievement, achievement_source = choose_field([
        (final_fields.get("achievement"), "dify.final_fields.achievement", confidence.get("achievement")),
        (row.get("申报理由"), "excel.申报理由", "medium"),
    ])
    result_row = {
        "序号": index,
        "奖项名称": row.get("申报项目", ""),
        "主体": subject,
        "所属BU": row.get("所属BU", ""),
        "所属PL": row.get("所属PL", ""),
        "提报人": submitter,
        "团队负责人": team_leader,
        "事迹": achievement,
    }
    field_sources = {
        "奖项名称": {"source": "excel.申报项目", "confidence": "high", "value": result_row["奖项名称"]},
        "主体": subject_source,
        "所属BU": {"source": "excel.所属BU", "confidence": "high", "value": result_row["所属BU"]},
        "所属PL": {"source": "excel.所属PL", "confidence": "high", "value": result_row["所属PL"]},
        "提报人": submitter_source,
        "团队负责人": team_leader_source,
        "事迹": achievement_source,
    }
    return result_row, field_sources


def calculate_score(review_json, award_config):
    weights = award_config.get("weights", DEFAULT_AWARD_CONFIG_DATA["default"]["weights"])
    evidence = review_json.get("evidence_grades", {}) if isinstance(review_json, dict) else {}
    total_weight = sum(float(value) for value in weights.values()) or 1
    dimensions = {}
    score = 0
    for dimension, weight in weights.items():
        item = evidence.get(dimension, {}) if isinstance(evidence.get(dimension, {}), dict) else {}
        grade = str(item.get("grade", "missing")).strip().lower() or "missing"
        points = GRADE_POINTS.get(grade, 0)
        weighted = points * float(weight) / total_weight
        dimensions[dimension] = {
            "grade": grade,
            "points": points,
            "weight": float(weight),
            "weighted_score": round(weighted, 2),
            "reason": item.get("reason", ""),
            "evidence": item.get("evidence", ""),
        }
        score += weighted

    risk_flags = review_json.get("risk_flags", []) if isinstance(review_json, dict) else []
    risk_penalty = min(
        len(risk_flags) * float(award_config.get("risk_penalty_per_flag", 0)),
        float(award_config.get("max_risk_penalty", 0)),
    )
    final_score = max(0, score - risk_penalty)
    return {
        "score": round(final_score, 2),
        "raw_score": round(score, 2),
        "risk_penalty": round(risk_penalty, 2),
        "dimensions": dimensions,
    }


def normalize_match_text(value):
    text = str(value or "").lower()
    return re.sub(r"[\s\-_—–&＋+()/（）·、，,。.:：;；]+", "", text)


def priority_subject_org(value):
    text = re.split(r"[\s\-_—–]+", str(value or "").strip(), maxsplit=1)[0]
    return normalize_match_text(text)


def leadership_priority_score(rank, total_count):
    if not rank or not total_count:
        return 0
    return round(max(0, (total_count - rank + 1) / total_count * 100), 2)


def leadership_source_label(path):
    stem = Path(path).stem
    if "汉霖" in stem:
        return "复宏汉霖"
    if "复星健康" in stem:
        return "复星健康"
    return stem.replace("优先级", "").strip() or stem


def build_fixed_leadership_priorities():
    rows = []
    for rule_index, rule in enumerate(FIXED_LEADERSHIP_PRIORITY_RULES, start=1):
        award_name = rule["award_name"]
        source_label = rule["source_label"]
        subjects = rule["subjects"]
        total_count = len(subjects)
        for rank, subject_item in enumerate(subjects, start=1):
            if isinstance(subject_item, dict):
                subject = subject_item["subject"]
                match_subjects = [subject] + list(subject_item.get("aliases", []))
            else:
                subject = subject_item
                match_subjects = [subject]
            for match_subject in match_subjects:
                rows.append({
                    "award_name": award_name,
                    "subject": subject,
                    "project": "",
                    "rank": rank,
                    "rank_explicit": True,
                    "source_row": rank,
                    "source_sheet": f"本地固定优先级-{rule_index}",
                    "source_path": "local:fixed_priority",
                    "source_label": source_label,
                    "priority_note": f"结合{source_label}内部战略优先级。",
                    "award_norm": normalize_match_text(award_name),
                    "subject_norm": normalize_match_text(match_subject),
                    "subject_org_norm": priority_subject_org(match_subject),
                    "project_norm": "",
                    "source_label_norm": normalize_match_text(source_label),
                    "total_count": total_count,
                    "priority_score": leadership_priority_score(rank, total_count),
                    "normal_weight": FIXED_NORMAL_REVIEW_WEIGHT,
                    "leadership_weight": FIXED_LEADERSHIP_PRIORITY_WEIGHT,
                })
    return rows


def priority_header_columns(ws):
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        headers = [cell_to_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        award_col = subject_col = project_col = rank_col = None
        for index, header in enumerate(headers, start=1):
            if not header:
                continue
            if award_col is None and "申报奖项" in header:
                award_col = index
            if subject_col is None and (
                "申报主体" in header
                or "申报部门" in header
                or "团队名称" in header
                or "Candidate" in header
            ):
                subject_col = index
            if project_col is None and (
                "项目概况" in header
                or "申报理由" in header
                or "事迹" in header
                or "Reasons" in header
            ):
                project_col = index
            if rank_col is None and ("排序" in header or "优先级" in header or "Proposal" in header):
                rank_col = index
        if award_col and subject_col and rank_col:
            return {
                "header_row": row_idx,
                "award_col": award_col,
                "subject_col": subject_col,
                "project_col": project_col,
                "rank_col": rank_col,
            }
    return None


def load_leadership_priority_file(path):
    if not path or not path.exists():
        return []

    wb = load_workbook(path, data_only=True)
    rows = []
    source_label = leadership_source_label(path)
    for ws in wb.worksheets:
        columns = priority_header_columns(ws)
        if not columns:
            continue
        current_award = ""
        for row_idx in range(columns["header_row"] + 1, ws.max_row + 1):
            award_name = cell_to_text(ws.cell(row_idx, columns["award_col"]).value) or current_award
            subject = cell_to_text(ws.cell(row_idx, columns["subject_col"]).value)
            project = cell_to_text(ws.cell(row_idx, columns["project_col"]).value) if columns["project_col"] else ""
            rank_text = cell_to_text(ws.cell(row_idx, columns["rank_col"]).value)
            if award_name:
                current_award = award_name
            if not current_award or not subject:
                continue
            if "签字" in subject or "Signature" in subject:
                continue
            rank = None
            try:
                if rank_text:
                    rank = int(float(rank_text))
            except ValueError:
                continue
            rows.append({
                "award_name": current_award,
                "subject": subject,
                "project": project,
                "rank": rank,
                "rank_explicit": rank is not None,
                "source_row": row_idx,
                "source_sheet": ws.title,
                "source_path": str(path),
                "source_label": source_label,
                "priority_note": f"结合{source_label}内部战略优先级。",
                "award_norm": normalize_match_text(current_award),
                "subject_norm": normalize_match_text(subject),
                "subject_org_norm": priority_subject_org(subject),
                "project_norm": normalize_match_text(project),
                "source_label_norm": normalize_match_text(source_label),
            })

    grouped = {}
    for item in rows:
        key = (item["source_label"], item["award_name"])
        grouped.setdefault(key, []).append(item)
    rows = []
    for group_items in grouped.values():
        explicit_ranks = [item["rank"] for item in group_items if item["rank_explicit"]]
        if not explicit_ranks:
            continue
        next_rank = max(explicit_ranks) + 1
        for item in group_items:
            if not item["rank_explicit"]:
                item["rank"] = next_rank
                next_rank += 1
            rows.append(item)
    counts = {}
    for item in rows:
        key = (item["source_label"], item["award_name"])
        counts[key] = max(counts.get(key, 0), item["rank"])
    for item in rows:
        item["total_count"] = counts.get((item["source_label"], item["award_name"]), item["rank"])
        item["priority_score"] = leadership_priority_score(item["rank"], item["total_count"])
    return rows


def load_leadership_priorities(paths, enabled=None):
    if enabled is None:
        enabled = ENABLE_LEADERSHIP_PRIORITY
    if not enabled:
        return []
    priorities = build_fixed_leadership_priorities()
    if not paths:
        return priorities
    if isinstance(paths, (str, Path)):
        paths = [paths]
    for path in paths:
        priorities.extend(load_leadership_priority_file(Path(path)))
    return priorities


def leadership_match_score(priority, row):
    subject_norm = normalize_match_text(row.get("申报主体", ""))
    blob_norm = normalize_match_text(json.dumps(row, ensure_ascii=False))
    priority_subject = priority["subject_norm"]
    source_label_norm = priority.get("source_label_norm", "")
    subject_org_norm = priority.get("subject_org_norm", "")
    source_label_present = bool(source_label_norm and len(source_label_norm) >= 3 and source_label_norm in blob_norm)
    subject_org_present = bool(subject_org_norm and len(subject_org_norm) >= 3 and subject_org_norm in blob_norm)
    source_present = source_label_present or subject_org_present
    subject_related = False
    score = 0

    if priority_subject and subject_norm:
        if priority_subject == subject_norm:
            score += 120
            subject_related = True
        elif priority_subject in subject_norm or subject_norm in priority_subject:
            score += 90
            subject_related = True
        elif priority_subject in blob_norm:
            score += 80
            subject_related = True
        else:
            priority_tail = priority_subject
            subject_tail = subject_norm
            for alias in ("复宏汉霖", "汉霖", "复星健康", priority.get("source_label_norm", "")):
                if alias:
                    priority_tail = priority_tail.replace(alias, "")
                    subject_tail = subject_tail.replace(alias, "")
            if source_present and priority_tail and subject_tail and (priority_tail in subject_tail or subject_tail in priority_tail):
                score += 70
                subject_related = True

    if not subject_related and not source_present:
        return 0

    if priority["project_norm"] and blob_norm:
        matcher = difflib.SequenceMatcher(None, priority["project_norm"], blob_norm)
        longest = matcher.find_longest_match(0, len(priority["project_norm"]), 0, len(blob_norm)).size
        score += min(60, longest * 3)
        score += matcher.ratio() * 30

    if not subject_related and not subject_org_present and score < 70:
        return 0

    return score


def match_leadership_priority(row, award_name, priorities):
    default = {
        "applied": False,
        "base_score": 0,
        "adjustment_score": 0,
        "normal_weight": 1.0,
        "leadership_weight": 0.0,
    }
    if not priorities:
        return default

    award_norm = normalize_match_text(award_name)
    same_award = [
        item
        for item in priorities
        if item["award_norm"] and (item["award_norm"] in award_norm or award_norm in item["award_norm"])
    ]
    if not same_award:
        return default

    scored = sorted(
        ((leadership_match_score(item, row), item) for item in same_award),
        key=lambda pair: pair[0],
        reverse=True,
    )
    match_score, priority = scored[0]
    if match_score < 35:
        return default

    normal_weight = float(priority.get("normal_weight", NORMAL_REVIEW_WEIGHT))
    leadership_weight = float(priority.get("leadership_weight", LEADERSHIP_PRIORITY_WEIGHT))
    return {
        "applied": True,
        "award_name": priority["award_name"],
        "subject": priority["subject"],
        "project": priority["project"],
        "source_label": priority.get("source_label", ""),
        "source_path": priority.get("source_path", ""),
        "source_sheet": priority.get("source_sheet", ""),
        "priority_note": priority.get("priority_note", ""),
        "proposal_rank": priority["rank"],
        "proposal_count": priority["total_count"],
        "priority_score": priority["priority_score"],
        "match_score": round(match_score, 2),
        "source_row": priority["source_row"],
        "normal_weight": normal_weight,
        "leadership_weight": leadership_weight,
    }


def apply_leadership_priority(base_score, priority):
    priority = copy.deepcopy(priority)
    priority["base_score"] = round(base_score, 2)
    if not priority.get("applied"):
        priority["adjustment_score"] = round(base_score, 2)
        return round(base_score, 2), priority
    normal_weight = float(priority.get("normal_weight", NORMAL_REVIEW_WEIGHT))
    leadership_weight = float(priority.get("leadership_weight", LEADERSHIP_PRIORITY_WEIGHT))
    adjustment_score = (
        float(priority.get("priority_score", 0)) * leadership_weight
        + float(base_score) * normal_weight
    )
    priority["adjustment_score"] = round(adjustment_score, 2)
    return round(base_score, 2), priority


def to_plain_text(value):
    if isinstance(value, dict):
        return first_non_empty(
            value.get("text"),
            value.get("risk"),
            value.get("flag"),
            value.get("description"),
            value.get("reason"),
            value.get("message"),
            json.dumps(value, ensure_ascii=False),
        )
    if isinstance(value, (list, tuple, set)):
        parts = []
        for item in value:
            text = to_plain_text(item)
            if text:
                parts.append(text)
        return "；".join(parts)
    return str(value or "").strip()


def needs_manual_review(review_json, workflow_status):
    if workflow_status != "succeeded":
        return True
    evidence = review_json.get("evidence_grades", {}) if isinstance(review_json, dict) else {}
    grades = [
        str((evidence.get(key) or {}).get("grade", "missing")).strip().lower()
        for key in DEFAULT_AWARD_CONFIG_DATA["default"]["weights"]
    ]
    if grades and all(grade == "missing" for grade in grades):
        return True
    risk_flags = review_json.get("risk_flags", []) if isinstance(review_json, dict) else []
    risk_text = " ".join(to_plain_text(item) for item in (risk_flags or []))
    blocking_keywords = ["一票否决", "风控", "合规", "重大", "完全缺失", "无具体申报理由", "无法进行任何有效评估"]
    return any(keyword in risk_text for keyword in blocking_keywords)


def tie_break_score(row, award_config):
    text = json.dumps(row, ensure_ascii=False)
    return sum(1 for keyword in award_config.get("tie_break_keywords", []) if keyword and keyword in text)


def completion_fields(result_row, final_fields):
    field_map = {
        "subject": "主体",
        "submitter": "提报人",
        "team_leader": "团队负责人",
        "team_members": "团队成员",
        "achievement": "事迹",
    }
    fields = []
    for header in TARGET_HEADERS:
        if header == "排名理由":
            continue
        if header != "序号" and str(result_row.get(header, "")).strip() in PLACEHOLDER_VALUES:
            fields.append(header)
    for field in (final_fields.get("needs_completion", []) if isinstance(final_fields, dict) else []):
        header = field_map.get(field, field)
        if header not in TARGET_HEADERS:
            continue
        if header not in fields:
            fields.append(header)
    return fields


def apply_leadership_slot_adjustment(sorted_entries):
    adjusted = list(sorted_entries)
    grouped_slots = {}
    for index, entry in enumerate(adjusted):
        priority = entry.get("leadership_priority", {})
        if priority.get("applied"):
            key = priority.get("source_label") or priority.get("source_path") or "default"
            grouped_slots.setdefault(key, []).append(index)

    for slots in grouped_slots.values():
        if len(slots) <= 1:
            continue
        prioritized_entries = [adjusted[index] for index in slots]
        prioritized_entries = sorted(
            prioritized_entries,
            key=lambda entry: (
                -float(entry.get("leadership_priority", {}).get("adjustment_score", 0)),
                entry.get("normal_award_rank", 999999),
                entry["excel_row"],
            ),
        )
        for slot_index, entry in zip(slots, prioritized_entries):
            adjusted[slot_index] = entry
    return adjusted


def rank_candidates(entries, config, dry_run=False):
    grouped = {}
    for entry in entries:
        grouped.setdefault(entry["award_name"], []).append(entry)

    for award_name, award_entries in grouped.items():
        award_config = get_award_config(config, award_name)
        quota = int(award_config.get("quota", 2) or 2)
        for entry in award_entries:
            entry["recommended_quota"] = quota
            entry["award_total_count"] = len(award_entries)

        sorted_entries = sorted(
            award_entries,
            key=lambda item: (
                -item["internal_score"],
                -item["tie_break_score"],
                item["manual_review_required"],
                item["excel_row"],
            ),
        )
        for normal_rank, entry in enumerate(sorted_entries, start=1):
            entry["normal_award_rank"] = normal_rank

        sorted_entries = apply_leadership_slot_adjustment(sorted_entries)
        eligible_count = 0
        for rank, entry in enumerate(sorted_entries, start=1):
            entry["award_rank"] = rank
            leadership_priority = entry.get("leadership_priority", {})
            if leadership_priority.get("applied"):
                leadership_priority["normal_award_rank"] = entry.get("normal_award_rank", rank)
                leadership_priority["final_award_rank"] = rank
                leadership_priority["slot_adjusted"] = entry.get("normal_award_rank", rank) != rank
            if dry_run:
                entry["recommendation_status"] = "dry_run"
                eligible_count += 1
                continue
            if entry["manual_review_required"]:
                entry["recommendation_status"] = "needs_review"
            elif eligible_count < quota:
                entry["recommendation_status"] = "recommended"
                eligible_count += 1
            else:
                entry["recommendation_status"] = "not_selected"

    return entries


DIMENSION_LABELS = {
    "rule_match": "规则匹配",
    "quantitative": "量化结果",
    "value_impact": "价值影响",
    "innovation": "创新性",
    "strategy_align": "战略契合",
}


def match_level(score):
    if score >= 80:
        return "高"
    if score >= 60:
        return "较高"
    if score >= 40:
        return "中等"
    if score > 0:
        return "偏低"
    return "缺失"


def build_ranking_reason(entry):
    score_detail = entry.get("score_detail", {})
    dimensions = score_detail.get("dimensions", {})
    strong_or_medium = []
    weak_or_missing = []
    for key, label in DIMENSION_LABELS.items():
        grade = (dimensions.get(key, {}) or {}).get("grade", "missing")
        reason = (dimensions.get(key, {}) or {}).get("reason", "")
        if grade in {"strong", "medium"}:
            strong_or_medium.append(label)
        elif reason:
            weak_or_missing.append(label)

    parts = [
        f"本奖项排名第{entry.get('award_rank', '')}位。综合匹配度{match_level(entry.get('internal_score', 0))}。"
    ]
    if strong_or_medium:
        parts.append(f"主要支撑来自{('、').join(strong_or_medium[:3])}。")
    else:
        parts.append("申报材料暂未形成足够的可评分证据。")
    points = ranking_reason_evidence_points(entry, 2)
    if points:
        parts.append(ensure_sentence(f"关键证据包括{('、').join(points[:2])}"))

    missing = entry.get("review_json", {}).get("missing_evidence", []) or []
    if weak_or_missing:
        parts.append(f"相对不足集中在{('、').join(weak_or_missing[:2])}。")
    stable_missing = [
        to_plain_text(item)
        for item in missing
        if not re.search(r"时间|未来|当前为20\d{2}|成立", to_plain_text(item))
    ]
    if stable_missing:
        missing_text = truncate_text(stable_missing[0], 80)
        parts.append(ensure_sentence(f"需补充/复核：{missing_text}"))
    if entry.get("recommendation_status") == "needs_review":
        parts.append("因此需人工复核后再确认最终意见。")
    return truncate_text("".join(parts), 320)


def truncate_text(value, limit=260):
    text = str(value or "").strip()
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "..."


def ensure_sentence(text):
    value = str(text or "").strip()
    if not value:
        return ""
    if value[-1] in "。！？；;.!?":
        return value
    return value + "。"


def remove_rank_one_bad_sentences(text):
    sentences = re.findall(r"[^。！？；;]+[。！？；;]?", text)
    if not sentences:
        sentences = [text]
    kept = [sentence for sentence in sentences if not any(marker in sentence for marker in RANK_ONE_BAD_REASON_MARKERS)]
    return "".join(kept).strip()


def remove_leading_rank_prefix(text):
    for pattern in (
        r"^\s*(?:该候选人)?(?:在本奖项中|在本批次评选中)?排名第[一二三四五六七八九十\d]+[位名]?"
        r"(?:，共[一二三四五六七八九十\d]+位候选人)?[，。；;]?",
        r"^\s*本奖项排名第[一二三四五六七八九十\d]+[位名]?[，。；;]?",
    ):
        text = re.sub(pattern, "", text).strip()
    return text


def normalize_ranking_reason(reason, rank):
    text = str(reason or "").strip()
    if not text:
        return ""
    if str(rank).strip() == "1":
        text = remove_rank_one_bad_sentences(text)
        for pattern in (
            r"[^。！？；;]*(?:与|和|同|较|比|相比|相较于|相较|相对)[^。！？；;]{0,20}(?:排名)?(?:更高|更前|靠前|前列|前两位|第一|领先|更高排名)(?:的)?(?:候选人|团队|项目|对象|者)?(?:相比|比较)?[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*(?:排名)?(?:更高|更前|靠前|前列|前两位|第一|领先|更高排名)(?:的)?(?:候选人|团队|项目|对象|者)[^。！？；;]*(?:相比|比较|优势|不足|差距)[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*与排名更前的候选人相比[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*与排名更高的候选人相比[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*与排名前列的候选人相比[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*与排名前两位的候选人相比[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*与排名第一的候选人相比[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*与排名靠前的候选人相比[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*相较于排名更高的候选人[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*相较于排名前列的候选人[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*相较于排名靠前的候选人[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*略逊[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*略逊一筹[^。！？；;]*[。！？；;]?",
            r"[^。！？；;]*亮点不够突出[^。！？；;]*[。！？；;]?",
        ):
            text = re.sub(pattern, "", text).strip()
    text = remove_leading_rank_prefix(text)
    text = re.sub(r"位居当前名次[，。；;]?", "", text).strip()
    prefix = f"本奖项排名第{rank}位。"
    if text.startswith(prefix):
        return truncate_text(text, 260)
    return truncate_text(prefix + text, 260)


def ranking_reason_body(reason_json):
    if not isinstance(reason_json, dict):
        return ""
    return first_non_empty(
        reason_json.get("reason_body"),
        reason_json.get("ranking_reason_body"),
        reason_json.get("core_reason"),
        reason_json.get("ranking_reason"),
        reason_json.get("reason"),
    )


def grade_labels(entry):
    dimensions = entry.get("score_detail", {}).get("dimensions", {})
    return {
        label: (dimensions.get(key, {}) or {}).get("grade", "missing")
        for key, label in DIMENSION_LABELS.items()
    }


def key_strengths(entry, limit=3):
    strengths = []
    dimensions = entry.get("score_detail", {}).get("dimensions", {})
    for key, label in DIMENSION_LABELS.items():
        item = dimensions.get(key, {}) or {}
        if item.get("grade") in {"strong", "medium"}:
            evidence = to_plain_text(item.get("evidence") or item.get("reason") or label)
            strengths.append(f"{label}: {truncate_text(evidence, 80)}")
    return strengths[:limit]


def key_gaps(entry, limit=2):
    gaps = entry.get("review_json", {}).get("missing_evidence", []) or []
    if gaps:
        return [truncate_text(to_plain_text(item), 90) for item in gaps[:limit]]
    dimensions = entry.get("score_detail", {}).get("dimensions", {})
    for key, label in DIMENSION_LABELS.items():
        item = dimensions.get(key, {}) or {}
        if item.get("grade") in {"weak", "missing"} and item.get("reason"):
            gaps.append(f"{label}: {item.get('reason')}")
    return [truncate_text(item, 90) for item in gaps[:limit]]


def unique_text_items(items, limit, text_limit=120):
    result = []
    seen = set()
    for item in items:
        text = re.sub(r"\s+", " ", to_plain_text(item)).strip(" -:：，,；;")
        if not text:
            continue
        text = truncate_text(text, text_limit)
        if text in seen:
            continue
        seen.add(text)
        result.append(text)
        if len(result) >= limit:
            break
    return result


def dimension_grade_aliases(entry):
    dimensions = entry.get("score_detail", {}).get("dimensions", {})
    grades = {}
    for key, label in DIMENSION_LABELS.items():
        grade = (dimensions.get(key, {}) or {}).get("grade", "missing")
        grades[key] = grade
        grades[label] = grade
    return grades


def summary_achievement(entry):
    fields = entry.get("result_row", {})
    raw_values = (entry.get("record", {}) or {}).get("values", {}) or {}
    final_fields = entry.get("final_fields_json", {}) or {}
    if not isinstance(final_fields, dict):
        final_fields = {}
    placeholders = PLACEHOLDER_VALUES | {"详见附件", "详见原文", "详见申报材料", "详见申报理由"}
    for value in (
        fields.get("事迹", ""),
        final_fields.get("事迹", ""),
        final_fields.get("achievement", ""),
        raw_values.get("申报理由", ""),
    ):
        text = str(value or "").strip()
        if text and text not in placeholders and not text.startswith("详见"):
            return text
    return ""


def evidence_keywords(entry, limit=8):
    candidates = []
    dimensions = entry.get("score_detail", {}).get("dimensions", {})
    for key in DIMENSION_LABELS:
        item = dimensions.get(key, {}) or {}
        for value in (item.get("evidence", ""), item.get("reason", "")):
            for part in re.split(r"[。；;\n\r]+", str(value or "")):
                candidates.append(part)
    for rule in entry.get("review_json", {}).get("matched_rules", []) or []:
        candidates.append(rule)
    if not candidates:
        candidates.append(summary_achievement(entry))
    return unique_text_items(candidates, limit, 100)


def ranking_reason_evidence_points(entry, limit=3):
    achievement = re.sub(r"[（(]\d+[）)]", "。", summary_achievement(entry))
    achievement_parts = []
    for part in re.split(r"[。；;\n\r]+", achievement):
        text = str(part or "").strip()
        if not text:
            continue
        if len(text) > 70:
            achievement_parts.extend(item for item in re.split(r"[，,]+", text) if str(item or "").strip())
        else:
            achievement_parts.append(text)
    outcome_pattern = r"\d|%|倍|提升|压缩|增长|获批|收入|利润|成本|效率|第|首|突破|全球|节省"
    preferred_parts = [part for part in achievement_parts if re.search(outcome_pattern, part)]
    candidates = preferred_parts + achievement_parts + evidence_keywords(entry, limit)
    return unique_text_items(candidates, limit, 70)


def build_candidate_summary(entry):
    fields = entry.get("result_row", {})
    support = key_strengths(entry, 5)
    gaps = key_gaps(entry, 3)
    keywords = evidence_keywords(entry)
    matched_rules = unique_text_items(entry.get("review_json", {}).get("matched_rules", []) or [], 8, 140)
    risk_flags = unique_text_items(entry.get("review_json", {}).get("risk_flags", []) or [], 3, 120)
    internal_score = entry.get("internal_score", 0)
    return {
        "candidate_id": entry.get("candidate_id", ""),
        "rank": entry.get("award_rank", ""),
        "total_count": entry.get("award_total_count", ""),
        "subject": fields.get("主体", ""),
        "bu": fields.get("所属BU", ""),
        "pl": fields.get("所属PL", ""),
        "submitter": fields.get("提报人", ""),
        "team_leader": fields.get("团队负责人", ""),
        "team_members": fields.get("团队成员", ""),
        "achievement": truncate_text(summary_achievement(entry), 700),
        "match_level": match_level(internal_score),
        "score": internal_score,
        "internal_score": internal_score,
        "normal_review_score": entry.get("normal_review_score", internal_score),
        "leadership_priority": entry.get("leadership_priority", {}),
        "evidence_grades": grade_labels(entry),
        "evidence_grade_aliases": dimension_grade_aliases(entry),
        "evidence_keywords": keywords,
        "matched_rules": matched_rules,
        "main_evidence": keywords,
        "key_support": support,
        "key_gap": gaps,
        "key_gaps": gaps,
        "missing_evidence": gaps,
        "risk_flags": risk_flags,
        "manual_review_note": entry.get("error", ""),
        "recommendation_status": entry.get("recommendation_status", ""),
        "manual_review_required": entry.get("manual_review_required", False),
    }


def build_award_ranking_summary(award_entries):
    return [
        {
            "rank": entry.get("award_rank", ""),
            "candidate_id": entry.get("candidate_id", ""),
            "subject": entry.get("result_row", {}).get("主体", ""),
            "bu": entry.get("result_row", {}).get("所属BU", ""),
            "pl": entry.get("result_row", {}).get("所属PL", ""),
            "match_level": match_level(entry.get("internal_score", 0)),
            "score": entry.get("internal_score", 0),
            "internal_score": entry.get("internal_score", 0),
            "normal_review_score": entry.get("normal_review_score", entry.get("internal_score", 0)),
            "leadership_priority": entry.get("leadership_priority", {}),
            "evidence_keywords": evidence_keywords(entry, 4),
            "key_support": key_strengths(entry, 2),
            "key_strength": "；".join(key_strengths(entry, 2)),
            "key_gaps": key_gaps(entry, 2),
            "key_gap": "；".join(key_gaps(entry, 2)),
            "manual_review_required": entry.get("manual_review_required", False),
        }
        for entry in sorted(award_entries, key=lambda item: (item.get("award_rank") or 999999, item["excel_row"]))
    ]


def build_review_rule_summary(award_entries):
    rules = []
    for entry in award_entries:
        for rule in entry.get("review_json", {}).get("matched_rules", []) or []:
            text = truncate_text(rule, 180)
            if text and text not in rules:
                rules.append(text)
    if rules:
        return "\n".join(f"- {rule}" for rule in rules[:6])
    return f"奖项名称：{award_entries[0]['award_name']}。请基于候选人的证据等级、缺失证据和同奖项对比生成排名理由。"


def build_ranking_reason_inputs(entry, award_entries):
    return {
        "batch_id": entry.get("batch_id", ""),
        "award_name": entry.get("award_name", ""),
        "candidate_id": entry.get("candidate_id", ""),
        "rank": str(entry.get("award_rank", "")),
        "total_count": str(entry.get("award_total_count", "")),
        "candidate_summary_json": json.dumps(build_candidate_summary(entry), ensure_ascii=False),
        "award_ranking_summary_json": json.dumps(build_award_ranking_summary(award_entries), ensure_ascii=False),
        "review_rule_summary": build_review_rule_summary(award_entries),
    }


def add_leadership_priority_context(reason, entry):
    priority = entry.get("leadership_priority", {})
    if not priority.get("applied"):
        return reason
    text = str(reason or "").strip()
    note = priority.get("priority_note") or f"结合{priority.get('source_label', '内部')}战略优先级。"
    if not text or note in text:
        return text
    return truncate_text(text + note, 360)


def is_low_quality_ranking_reason(reason_body, entry):
    text = str(reason_body or "")
    if not text:
        return False
    markers = (
        "当前正文证据不足",
        "证据不足",
        "需人工结合原始材料复核",
        "无法提取有效事实",
        "evidence_keywords 为空",
        "achievement 为\"详见附件\"",
    )
    has_local_evidence = bool(key_strengths(entry, 1) or evidence_keywords(entry, 1) or summary_achievement(entry))
    return has_local_evidence and any(marker in text for marker in markers)


def generate_ranking_reasons(entries, base_url, api_key, user, timeout, dry_run=False):
    grouped = {}
    for entry in entries:
        grouped.setdefault(entry["award_name"], []).append(entry)

    for award_name, award_entries in grouped.items():
        for entry in sorted(award_entries, key=lambda item: (item.get("award_rank") or 999999, item["excel_row"])):
            fallback = add_leadership_priority_context(build_ranking_reason(entry), entry)
            entry["ranking_reason_source"] = "local_fallback"
            entry["ranking_reason_error"] = ""
            entry["ranking_reason_body"] = ""
            if dry_run or not api_key:
                entry["ranking_reason"] = fallback
                continue
            try:
                inputs = build_ranking_reason_inputs(entry, award_entries)
                _, outputs = call_dify_workflow(base_url, api_key, inputs, user, timeout)
                reason_json = parse_json_maybe(outputs.get("ranking_reason_json") or outputs.get("text") or outputs)
                reason_body = ranking_reason_body(reason_json)
                reason = normalize_ranking_reason(reason_body, entry.get("award_rank", ""))
                entry["ranking_reason_body"] = reason_body
                entry["ranking_reason_json"] = reason_json
                if reason and is_low_quality_ranking_reason(reason_body, entry):
                    entry["ranking_reason"] = fallback
                    entry["ranking_reason_source"] = "local_quality_repair"
                else:
                    entry["ranking_reason"] = add_leadership_priority_context(reason or fallback, entry)
                    entry["ranking_reason_source"] = "dify" if reason else "local_fallback"
            except Exception as exc:
                entry["ranking_reason"] = fallback
                entry["ranking_reason_body"] = ""
                entry["ranking_reason_error"] = str(exc)
                entry["ranking_reason_json"] = {}


def sorted_entries_for_result(entries):
    award_order = []
    grouped = {}
    for entry in entries:
        award_name = entry["award_name"]
        if award_name not in grouped:
            award_order.append(award_name)
            grouped[award_name] = []
        grouped[award_name].append(entry)

    sorted_entries = []
    for award_name in award_order:
        sorted_entries.extend(sorted(grouped[award_name], key=lambda item: (item.get("award_rank") or 999999, item["excel_row"])))
    return sorted_entries


def build_detail_row(entry):
    return {
        "candidate_id": entry["candidate_id"],
        "excel_row": entry["excel_row"],
        "batch_id": entry["batch_id"],
        "award_name": entry["award_name"],
        "award_type": entry["award_type"],
        "workflow_status": entry["workflow_status"],
        "recommendation_status": entry["recommendation_status"],
        "award_rank": entry["award_rank"],
        "recommended_quota": entry["recommended_quota"],
        "internal_score": entry["internal_score"],
        "normal_review_score": entry.get("normal_review_score", entry["internal_score"]),
        "leadership_priority_applied": entry.get("leadership_priority", {}).get("applied", False),
        "leadership_priority_score": entry.get("leadership_priority", {}).get("priority_score", ""),
        "leadership_adjustment_score": entry.get("leadership_priority", {}).get("adjustment_score", ""),
        "leadership_priority_rank": entry.get("leadership_priority", {}).get("proposal_rank", ""),
        "leadership_priority_subject": entry.get("leadership_priority", {}).get("subject", ""),
        "leadership_priority_source": entry.get("leadership_priority", {}).get("source_label", ""),
        "leadership_priority_match_score": entry.get("leadership_priority", {}).get("match_score", ""),
        "normal_award_rank": entry.get("normal_award_rank", ""),
        "leadership_slot_adjusted": entry.get("leadership_priority", {}).get("slot_adjusted", False),
        "manual_review_required": entry["manual_review_required"],
        "tie_break_score": entry["tie_break_score"],
        "completion_fields": json.dumps(entry["completion_fields"], ensure_ascii=False),
        "error": entry["error"],
        "missing_evidence": json.dumps(entry["review_json"].get("missing_evidence", []), ensure_ascii=False),
        "risk_flags": json.dumps(entry["review_json"].get("risk_flags", []), ensure_ascii=False),
        "explanation": entry["review_json"].get("explanation", ""),
        "ranking_reason": entry.get("ranking_reason", ""),
        "ranking_reason_body": entry.get("ranking_reason_body", ""),
        "ranking_reason_source": entry.get("ranking_reason_source", ""),
        "ranking_reason_error": entry.get("ranking_reason_error", ""),
        "ranking_reason_json": json.dumps(entry.get("ranking_reason_json", {}), ensure_ascii=False),
        "field_sources": json.dumps(entry["field_sources"], ensure_ascii=False),
        "score_detail": json.dumps(entry["score_detail"], ensure_ascii=False),
        "review_result_json": json.dumps(entry["review_json"] or entry["outputs"].get("review_result_json", ""), ensure_ascii=False),
        "final_fields_json": json.dumps(entry["final_fields_json"] or entry["outputs"].get("final_fields_json", ""), ensure_ascii=False),
        "raw_row_json": entry["inputs"]["raw_row_json"],
    }


def copy_row_style(ws, source_row, target_row):
    if source_row > ws.max_row:
        return
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)


def capture_row_style(ws, source_row):
    styles = []
    if source_row > ws.max_row:
        return styles
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        styles.append({
            "style": copy.copy(source._style) if source.has_style else None,
            "number_format": source.number_format,
            "alignment": copy.copy(source.alignment) if source.alignment else None,
        })
    return styles


def apply_captured_row_style(ws, target_row, styles):
    for col_idx, style in enumerate(styles, start=1):
        target = ws.cell(target_row, col_idx)
        if style.get("style"):
            target._style = copy.copy(style["style"])
        if style.get("number_format"):
            target.number_format = style["number_format"]
        if style.get("alignment"):
            target.alignment = copy.copy(style["alignment"])


def ensure_result_headers(ws):
    for col_idx, header in enumerate(TARGET_HEADERS, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        if col_idx > 1 and not cell.has_style:
            previous = ws.cell(1, col_idx - 1)
            if previous.has_style:
                cell._style = copy.copy(previous._style)
    reason_col = len(TARGET_HEADERS)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions[get_column_letter(reason_col)].width = 60
    ws.cell(1, reason_col).alignment = Alignment(wrap_text=True, vertical="center")


def merge_award_name_cells(ws):
    if ws.max_row <= 2:
        return
    start_row = 2
    current_award = ws.cell(start_row, 1).value
    for row_idx in range(3, ws.max_row + 2):
        award_name = ws.cell(row_idx, 1).value if row_idx <= ws.max_row else None
        if award_name == current_award:
            continue
        end_row = row_idx - 1
        if current_award and end_row > start_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            cell = ws.cell(start_row, 1)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        start_row = row_idx
        current_award = award_name


def clear_data_rows(ws):
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= 2:
            ws.unmerge_cells(str(merged_range))
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for key in list(ws._cells):
        if key[0] >= 2:
            del ws._cells[key]
    for row_idx in list(ws.row_dimensions):
        if row_idx >= 2:
            del ws.row_dimensions[row_idx]


def replace_with_empty_result_sheet(wb, old_ws):
    sheet_index = wb.worksheets.index(old_ws)
    header_style = capture_row_style(old_ws, 1)
    column_widths = {
        col_idx: old_ws.column_dimensions[get_column_letter(col_idx)].width
        for col_idx in range(1, len(TARGET_HEADERS) + 1)
    }
    ws = wb.create_sheet("评选结果_clean", sheet_index)
    wb.remove(old_ws)
    ws.title = "评选结果"
    ws.append(TARGET_HEADERS)
    if header_style:
        apply_captured_row_style(ws, 1, header_style)
    for col_idx, width in column_widths.items():
        if width:
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    return ws


def write_results_xlsx(template_path, output_path, result_rows, detail_rows):
    if template_path.exists():
        wb = load_workbook(template_path)
        ws = wb.active
        ws.title = "评选结果"
        ensure_result_headers(ws)
        if result_rows:
            row_style = capture_row_style(ws, 2)
            clear_data_rows(ws)
        else:
            ws = replace_with_empty_result_sheet(wb, ws)
            row_style = []
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "评选结果"
        ws.append(TARGET_HEADERS)
        ensure_result_headers(ws)
        row_style = []

    for row_data in result_rows:
        row_idx = ws.max_row + 1
        if row_style:
            apply_captured_row_style(ws, row_idx, row_style)
        else:
            copy_row_style(ws, 2, row_idx)
        for col_idx, header in enumerate(TARGET_HEADERS, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = row_data.get(header, "")
            if header == "排名理由":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    merge_award_name_cells(ws)

    if "评审明细" in wb.sheetnames:
        del wb["评审明细"]
    detail_ws = wb.create_sheet("评审明细")
    detail_headers = [
        "candidate_id",
        "excel_row",
        "batch_id",
        "award_name",
        "award_type",
        "workflow_status",
        "recommendation_status",
        "award_rank",
        "recommended_quota",
        "internal_score",
        "normal_review_score",
        "leadership_priority_applied",
        "leadership_priority_score",
        "leadership_adjustment_score",
        "leadership_priority_rank",
        "leadership_priority_subject",
        "leadership_priority_source",
        "leadership_priority_match_score",
        "normal_award_rank",
        "leadership_slot_adjusted",
        "manual_review_required",
        "tie_break_score",
        "completion_fields",
        "error",
        "missing_evidence",
        "risk_flags",
        "explanation",
        "ranking_reason",
        "ranking_reason_body",
        "ranking_reason_source",
        "ranking_reason_error",
        "ranking_reason_json",
        "field_sources",
        "score_detail",
        "review_result_json",
        "final_fields_json",
        "raw_row_json",
    ]
    detail_ws.append(detail_headers)
    for row_data in detail_rows:
        detail_ws.append([row_data.get(header, "") for header in detail_headers])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_internal_pack(path, entries):
    with path.open("w", encoding="utf-8") as file:
        for entry in entries:
            payload = {
                "candidate_id": entry["candidate_id"],
                "excel_row": entry["excel_row"],
                "batch_id": entry["batch_id"],
                "award_name": entry["award_name"],
                "award_type": entry["award_type"],
                "workflow_status": entry.get("workflow_status", ""),
                "workflow_error": entry.get("error", ""),
                "recommendation": {
                    "status": entry["recommendation_status"],
                    "rank": entry["award_rank"],
                    "quota": entry["recommended_quota"],
                    "manual_review_required": entry["manual_review_required"],
                    "ranking_reason": entry.get("ranking_reason", ""),
                    "ranking_reason_source": entry.get("ranking_reason_source", ""),
                    "ranking_reason_error": entry.get("ranking_reason_error", ""),
                },
                "scoring": {
                    "internal_score": entry["internal_score"],
                    "normal_review_score": entry.get("normal_review_score", entry["internal_score"]),
                    "leadership_priority": entry.get("leadership_priority", {}),
                    "tie_break_score": entry["tie_break_score"],
                    "detail": entry["score_detail"],
                },
                "evidence": {
                    "grades": entry["review_json"].get("evidence_grades", {}),
                    "matched_rules": entry["review_json"].get("matched_rules", []),
                    "missing_evidence": entry["review_json"].get("missing_evidence", []),
                    "risk_flags": entry["review_json"].get("risk_flags", []),
                    "explanation": entry["review_json"].get("explanation", ""),
                },
                "final_result_fields": entry["result_row"],
                "ranking_reason": entry.get("ranking_reason", ""),
                "ranking_reason_body": entry.get("ranking_reason_body", ""),
                "ranking_reason_json": entry.get("ranking_reason_json", {}),
                "field_sources": entry["field_sources"],
                "completion_fields": entry["completion_fields"],
                "raw_row": entry["record"]["values"],
            }
            file.write(json.dumps(payload, ensure_ascii=False) + "\n")


def write_completion_xlsx(path, entries):
    wb = Workbook()
    ws = wb.active
    ws.title = "待补充清单"
    headers = ["candidate_id", "奖项名称", "主体", "字段", "当前值", "推荐状态", "备注"]
    ws.append(headers)
    for entry in entries:
        for field in entry["completion_fields"]:
            ws.append([
                entry["candidate_id"],
                entry["award_name"],
                entry["result_row"].get("主体", ""),
                field,
                entry["result_row"].get(field, ""),
                entry["recommendation_status"],
                "请人工补充或复核",
            ])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def add_qa_check(report, check_id, passed, message, details=None):
    report["checks"].append({
        "id": check_id,
        "passed": bool(passed),
        "message": message,
        "details": details or {},
    })


def read_internal_pack(path):
    items = []
    with path.open(encoding="utf-8") as file:
        for line in file:
            if line.strip():
                items.append(json.loads(line))
    return items


def reconstructed_award_rows(ws):
    rows = []
    current_award = ""
    for row_idx in range(2, ws.max_row + 1):
        award = ws.cell(row_idx, 1).value
        if award:
            current_award = award
        rows.append({
            "row_idx": row_idx,
            "award_name": current_award,
            "rank": ws.cell(row_idx, 2).value,
            "ranking_reason": ws.cell(row_idx, len(TARGET_HEADERS)).value or "",
        })
    return rows


def award_groups_from_rows(rows):
    groups = []
    current = None
    for row in rows:
        if current is None or row["award_name"] != current["award_name"]:
            current = {"award_name": row["award_name"], "rows": []}
            groups.append(current)
        current["rows"].append(row)
    return groups


def run_quality_checks(xlsx_path, internal_pack_path, expected_rows, require_dify_reasons=True):
    report = {
        "schema_version": "review_batch_qa_v1",
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "xlsx_path": str(xlsx_path),
        "internal_pack_path": str(internal_pack_path),
        "expected_rows": expected_rows,
        "passed": False,
        "checks": [],
        "summary": {},
    }

    if not xlsx_path.exists():
        add_qa_check(report, "xlsx_exists", False, "结果 Excel 不存在")
        return report
    if not internal_pack_path.exists():
        add_qa_check(report, "internal_pack_exists", False, "内部评审包不存在")
        return report
    add_qa_check(report, "output_files_exist", True, "结果 Excel 和内部评审包均已生成")

    wb = load_workbook(xlsx_path, data_only=True)
    sheet_ok = "评选结果" in wb.sheetnames and "评审明细" in wb.sheetnames
    add_qa_check(report, "required_sheets", sheet_ok, "结果表包含评选结果和评审明细 sheet", {"sheets": wb.sheetnames})
    ws = wb["评选结果"] if "评选结果" in wb.sheetnames else wb.active

    headers = [ws.cell(1, col_idx).value for col_idx in range(1, ws.max_column + 1)]
    add_qa_check(report, "headers_match", headers == TARGET_HEADERS, "表头符合目标输出格式", {"headers": headers})
    data_rows = max(ws.max_row - 1, 0)
    add_qa_check(report, "row_count", data_rows == expected_rows, "结果行数等于输入候选行数", {
        "actual": data_rows,
        "expected": expected_rows,
    })
    add_qa_check(report, "column_count", ws.max_column == len(TARGET_HEADERS), "结果列数符合目标格式", {
        "actual": ws.max_column,
        "expected": len(TARGET_HEADERS),
    })

    rows = reconstructed_award_rows(ws)
    groups = award_groups_from_rows(rows)
    nonempty_reasons = sum(1 for row in rows if str(row["ranking_reason"]).strip())
    add_qa_check(report, "ranking_reasons_nonempty", nonempty_reasons == data_rows, "每行均有排名理由", {
        "actual": nonempty_reasons,
        "expected": data_rows,
    })

    merged_ranges = {str(merged_range) for merged_range in ws.merged_cells.ranges}
    merge_errors = []
    rank_errors = []
    prefix_errors = []
    rank_one_bad = []
    global_bad = []
    for group in groups:
        group_rows = group["rows"]
        if len(group_rows) > 1:
            expected_range = f"A{group_rows[0]['row_idx']}:A{group_rows[-1]['row_idx']}"
            if expected_range not in merged_ranges:
                merge_errors.append({"award_name": group["award_name"], "expected_range": expected_range})
        expected_ranks = list(range(1, len(group_rows) + 1))
        actual_ranks = []
        for row in group_rows:
            try:
                rank_value = int(row["rank"])
            except (TypeError, ValueError):
                rank_value = row["rank"]
            actual_ranks.append(rank_value)
            expected_prefix = f"本奖项排名第{rank_value}位。"
            reason = row["ranking_reason"]
            if reason and not reason.startswith(expected_prefix):
                prefix_errors.append({
                    "row": row["row_idx"],
                    "rank": rank_value,
                    "reason_preview": reason[:160],
                })
            if str(rank_value) == "1":
                matched = [marker for marker in RANK_ONE_BAD_REASON_MARKERS if marker in reason]
                if matched:
                    rank_one_bad.append({
                        "row": row["row_idx"],
                        "markers": matched,
                        "reason_preview": reason[:220],
                    })
            matched_global = [marker for marker in GLOBAL_BAD_REASON_MARKERS if marker in reason]
            if matched_global:
                global_bad.append({
                    "row": row["row_idx"],
                    "markers": matched_global,
                    "reason_preview": reason[:220],
                })
        if actual_ranks != expected_ranks:
            rank_errors.append({
                "award_name": group["award_name"],
                "actual": actual_ranks,
                "expected": expected_ranks,
            })

    add_qa_check(report, "award_name_merges", not merge_errors, "同奖项 A 列单元格已合并", {
        "merged_ranges": sorted(merged_ranges),
        "errors": merge_errors,
    })
    add_qa_check(report, "award_rank_sequence", not rank_errors, "每个奖项内排名从 1 连续递增", {"errors": rank_errors})
    add_qa_check(report, "ranking_reason_prefix", not prefix_errors, "排名理由前缀与 B 列排名一致", {"errors": prefix_errors})
    add_qa_check(report, "rank_one_forbidden_phrases", not rank_one_bad, "第一名不含更高/靠前候选人比较等禁用表达", {
        "errors": rank_one_bad,
    })
    add_qa_check(report, "global_broken_phrases", not global_bad, "排名理由不含残句", {"errors": global_bad})

    pack_items = read_internal_pack(internal_pack_path)
    workflow_errors = [
        {
            "candidate_id": item.get("candidate_id"),
            "workflow_status": item.get("workflow_status"),
            "workflow_error": item.get("workflow_error", ""),
        }
        for item in pack_items
        if item.get("workflow_status") not in {"succeeded", "dry_run"}
    ]
    reason_errors = [
        {
            "candidate_id": item.get("candidate_id"),
            "error": item.get("recommendation", {}).get("ranking_reason_error", ""),
        }
        for item in pack_items
        if item.get("recommendation", {}).get("ranking_reason_error")
    ]
    source_counts = {}
    for item in pack_items:
        source = item.get("recommendation", {}).get("ranking_reason_source", "")
        source_counts[source] = source_counts.get(source, 0) + 1
    dify_count = source_counts.get("dify", 0)
    allowed_reason_sources = {"dify", "leadership_priority_template", "local_quality_repair"}
    unexpected_reason_sources = [
        {
            "candidate_id": item.get("candidate_id"),
            "source": item.get("recommendation", {}).get("ranking_reason_source", ""),
        }
        for item in pack_items
        if item.get("recommendation", {}).get("ranking_reason_source", "") not in allowed_reason_sources
    ]

    add_qa_check(report, "internal_pack_row_count", len(pack_items) == expected_rows, "内部评审包行数等于输入候选行数", {
        "actual": len(pack_items),
        "expected": expected_rows,
    })
    add_qa_check(report, "workflow_status_succeeded", not workflow_errors, "评审 Workflow 全部成功", {
        "errors": workflow_errors,
    })
    add_qa_check(report, "ranking_reason_errors_empty", not reason_errors, "排名理由 Workflow 无错误", {
        "errors": reason_errors,
    })
    if require_dify_reasons:
        add_qa_check(report, "ranking_reason_source_allowed", not unexpected_reason_sources, "排名理由来自 Dify、内部优先级固定口径模板或本地质量修复", {
            "source_counts": source_counts,
            "dify_count": dify_count,
            "allowed_sources": sorted(allowed_reason_sources),
            "unexpected_sources": unexpected_reason_sources,
        })
    else:
        add_qa_check(report, "ranking_reason_source_recorded", bool(source_counts), "排名理由来源已记录", {
            "source_counts": source_counts,
        })

    report["summary"] = {
        "data_rows": data_rows,
        "nonempty_reasons": nonempty_reasons,
        "source_counts": source_counts,
        "checks_total": len(report["checks"]),
        "checks_failed": sum(1 for check in report["checks"] if not check["passed"]),
    }
    report["passed"] = all(check["passed"] for check in report["checks"])
    return report


def write_quality_report(path, report):
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def run_review_batch(config, event_sink=None, should_cancel=None):
    sink = event_sink or NullEventSink()
    cancelled = should_cancel or (lambda: False)

    base_url = config.dify_base_url.strip()
    api_key = config.dify_review_api_key.strip()
    ranking_reason_api_key = config.dify_ranking_reason_api_key.strip()
    user = config.dify_user.strip() or "review-batch"
    if not base_url:
        raise ValueError("缺少 DIFY_BASE_URL，请在 .env 中配置。")
    if not config.dry_run and not api_key:
        raise ValueError("缺少 DIFY_REVIEW_WORKFLOW_API_KEY，请在 .env 中配置。")

    records = read_excel_records(config.input_path)
    if config.award_filters:
        filters = [text.strip() for text in config.award_filters if text and text.strip()]
        records = [
            record
            for record in records
            if any(keyword in str(record["values"].get("申报项目", "")) for keyword in filters)
        ]
    if config.limit:
        records = records[: config.limit]
    sink.emit(
        "excel:loaded",
        message=f"载入 {len(records)} 个候选",
        progress=(0, len(records)),
        payload={"total": len(records)},
    )

    award_config = load_award_config(config.award_config_path, config.top_n)
    leadership_priority_paths = config.leadership_priority_paths or DEFAULT_LEADERSHIP_PRIORITIES
    leadership_priorities = load_leadership_priorities(
        leadership_priority_paths,
        enabled=config.enable_leadership_priority,
    )

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = config.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    jsonl_path = output_dir / f"review_results_{stamp}.jsonl"
    xlsx_path = output_dir / f"review_results_{stamp}.xlsx"
    internal_pack_path = output_dir / f"internal_review_pack_{stamp}.jsonl"
    completion_path = output_dir / f"待补充清单_{stamp}.xlsx"
    qa_report_path = output_dir / f"qa_report_{stamp}.json"

    candidate_entries = []
    with jsonl_path.open("w", encoding="utf-8") as jsonl:
        for idx, record in enumerate(records, start=1):
            if cancelled():
                raise RunCancelled()
            inputs = build_workflow_inputs(record)
            sink.emit(
                "candidate:started",
                message=f"第 {idx}/{len(records)} 行开始",
                progress=(idx - 1, len(records)),
                payload={"candidate_id": inputs["candidate_id"], "award_name": inputs["award_name"]},
            )
            workflow_status = "dry_run" if config.dry_run else "pending"
            error = ""
            outputs = {}
            response_body = {}
            review_json = {}
            final_fields_json = {}
            try:
                if not config.dry_run:
                    response_body, outputs = call_dify_workflow(base_url, api_key, inputs, user, config.timeout)
                    workflow_status = response_body.get("data", {}).get("status", "succeeded")
                    review_json = parse_json_maybe(outputs.get("review_result_json"))
                    final_fields_json = parse_json_maybe(outputs.get("final_fields_json"))
            except Exception as exc:
                workflow_status = "failed"
                error = str(exc)

            result_row, field_sources = build_result_row(idx, record, final_fields_json)
            current_award_config = get_award_config(award_config, inputs["award_name"])
            score_detail = calculate_score(review_json, current_award_config)
            leadership_priority = match_leadership_priority(record["values"], inputs["award_name"], leadership_priorities)
            internal_score, leadership_priority = apply_leadership_priority(score_detail["score"], leadership_priority)
            score_detail["leadership_priority"] = leadership_priority
            entry = {
                "candidate_id": inputs["candidate_id"],
                "excel_row": record["excel_row"],
                "batch_id": inputs["batch_id"],
                "award_name": inputs["award_name"],
                "award_type": inputs["award_type"],
                "workflow_status": workflow_status,
                "error": error,
                "outputs": outputs,
                "response_body": response_body,
                "inputs": inputs,
                "record": record,
                "review_json": review_json,
                "final_fields_json": final_fields_json,
                "result_row": result_row,
                "field_sources": field_sources,
                "score_detail": score_detail,
                "normal_review_score": score_detail["score"],
                "internal_score": internal_score,
                "leadership_priority": leadership_priority,
                "tie_break_score": tie_break_score(record["values"], current_award_config),
                "manual_review_required": False if config.dry_run else needs_manual_review(review_json, workflow_status),
                "completion_fields": completion_fields(result_row, final_fields_json),
                "recommendation_status": "pending",
                "award_rank": "",
                "recommended_quota": int(current_award_config.get("quota", config.top_n) or config.top_n),
            }
            candidate_entries.append(entry)
            jsonl.write(json.dumps({"inputs": inputs, "outputs": outputs, "response": response_body}, ensure_ascii=False) + "\n")
            jsonl.flush()
            print(f"[{idx}/{len(records)}] {inputs['candidate_id']} {inputs['award_name']} -> {workflow_status}", flush=True)
            if workflow_status in {"succeeded", "dry_run"}:
                sink.emit(
                    "candidate:reviewed",
                    message=f"第 {idx}/{len(records)} 行完成",
                    progress=(idx, len(records)),
                    payload={"candidate_id": inputs["candidate_id"], "workflow_status": workflow_status},
                )
            else:
                sink.emit(
                    "candidate:failed",
                    message=f"第 {idx}/{len(records)} 行失败",
                    level="warn",
                    progress=(idx, len(records)),
                    payload={"candidate_id": inputs["candidate_id"], "workflow_status": workflow_status, "error": error},
                )
            if not config.dry_run and config.sleep:
                time.sleep(config.sleep)

    if cancelled():
        raise RunCancelled()
    sink.emit("ranking:started", message="开始排序")
    rank_candidates(candidate_entries, award_config, dry_run=config.dry_run)
    sink.emit("ranking:done", message="排序完成")

    if cancelled():
        raise RunCancelled()
    sink.emit("reason:started", message="开始生成排名理由")
    generate_ranking_reasons(
        candidate_entries,
        base_url,
        ranking_reason_api_key,
        user,
        config.timeout,
        dry_run=config.dry_run,
    )
    sink.emit("reason:done", message="排名理由生成完成")

    if cancelled():
        raise RunCancelled()
    result_rows = []
    for entry in sorted_entries_for_result(candidate_entries):
        entry["result_row"][TARGET_HEADERS[1]] = entry["award_rank"]
        entry["result_row"][TARGET_HEADERS[-1]] = entry.get("ranking_reason", "")
        result_rows.append(copy.deepcopy(entry["result_row"]))
    detail_rows = [build_detail_row(entry) for entry in candidate_entries]
    sink.emit("export:started", message="开始导出结果文件")
    write_results_xlsx(config.template_path, xlsx_path, result_rows, detail_rows)
    sink.emit("artifact:created", payload={"artifact_type": "review_results_xlsx", "name": xlsx_path.name, "path": str(xlsx_path)})
    write_internal_pack(internal_pack_path, candidate_entries)
    sink.emit("artifact:created", payload={"artifact_type": "internal_review_pack", "name": internal_pack_path.name, "path": str(internal_pack_path)})
    write_completion_xlsx(completion_path, candidate_entries)
    sink.emit("artifact:created", payload={"artifact_type": "completion_xlsx", "name": completion_path.name, "path": str(completion_path)})

    sink.emit("qa:started", message="开始 QA 检查")
    qa_report = run_quality_checks(
        xlsx_path,
        internal_pack_path,
        expected_rows=len(records),
        require_dify_reasons=not config.dry_run,
    )
    write_quality_report(qa_report_path, qa_report)
    sink.emit("artifact:created", payload={"artifact_type": "qa_report", "name": qa_report_path.name, "path": str(qa_report_path)})
    failed_checks = [check["id"] for check in qa_report["checks"] if not check["passed"]]
    sink.emit(
        "qa:done",
        level="info" if qa_report["passed"] else "warn",
        payload={"passed": qa_report["passed"], "failed_checks": failed_checks},
    )

    award_counts = {}
    for entry in candidate_entries:
        award_counts[entry["award_name"]] = award_counts.get(entry["award_name"], 0) + 1
    return ReviewBatchResult(
        output_dir=output_dir,
        xlsx_path=xlsx_path,
        raw_jsonl_path=jsonl_path,
        internal_pack_path=internal_pack_path,
        completion_path=completion_path,
        qa_report_path=qa_report_path,
        qa_passed=qa_report["passed"],
        expected_rows=len(records),
        processed_rows=len(candidate_entries),
        award_counts=award_counts,
        qa_report=qa_report,
    )


def main():
    parser = argparse.ArgumentParser(description="批量调用 Dify 评优 Workflow，输出评选结果表。")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="输入 Excel 路径")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE), help="目标输出格式模板 Excel 路径")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="输出目录")
    parser.add_argument("--award-config", default=str(DEFAULT_AWARD_CONFIG), help="奖项名额和权重配置 JSON 路径")
    parser.add_argument("--leadership-priority", action="append", default=[], help="领导优先级 Excel 路径，可重复传入；默认使用代码内置固定优先级，显式传入时才额外读取 Excel")
    parser.add_argument("--disable-leadership-priority", action="store_true", help="关闭代码内置和显式传入的领导优先级，仅按正常评审分排序")
    parser.add_argument("--top-n", type=int, default=2, help="每个奖项默认拟推荐人数")
    parser.add_argument("--env-file", action="append", default=[], help="额外 .env 文件路径，可重复传入")
    parser.add_argument("--award-filter", action="append", default=[], help="仅处理申报项目包含该文本的行，可重复传入")
    parser.add_argument("--limit", type=int, default=0, help="仅处理前 N 行，0 表示全部")
    parser.add_argument("--sleep", type=float, default=0.2, help="每行调用后的暂停秒数")
    parser.add_argument("--timeout", type=int, default=120, help="Dify API 超时时间，秒")
    parser.add_argument("--dry-run", action="store_true", help="只生成 Workflow 输入，不调用 Dify")
    args = parser.parse_args()

    env_paths = [Path(path) for path in args.env_file] + DEFAULT_ENV_FILES
    load_env_files(env_paths)

    base_url = os.environ.get("DIFY_BASE_URL", "").strip()
    api_key = os.environ.get("DIFY_REVIEW_WORKFLOW_API_KEY", "").strip()
    ranking_reason_api_key = os.environ.get("DIFY_RANKING_REASON_WORKFLOW_API_KEY", "").strip()
    user = os.environ.get("DIFY_USER", "review-batch").strip() or "review-batch"
    if not base_url:
        raise SystemExit("缺少 DIFY_BASE_URL，请在 .env 中配置。")
    if not args.dry_run and not api_key:
        raise SystemExit("缺少 DIFY_REVIEW_WORKFLOW_API_KEY，请在 .env 中配置。")

    leadership_priority_paths = [Path(path) for path in args.leadership_priority] if args.leadership_priority else DEFAULT_LEADERSHIP_PRIORITIES
    config = ReviewBatchConfig(
        input_path=Path(args.input),
        output_dir=Path(args.output_dir),
        template_path=Path(args.template),
        award_config_path=Path(args.award_config),
        leadership_priority_paths=leadership_priority_paths,
        enable_leadership_priority=not args.disable_leadership_priority,
        top_n=args.top_n,
        award_filters=args.award_filter,
        limit=args.limit,
        sleep=args.sleep,
        timeout=args.timeout,
        dry_run=args.dry_run,
        dify_base_url=base_url,
        dify_review_api_key=api_key,
        dify_ranking_reason_api_key=ranking_reason_api_key,
        dify_user=user,
    )
    result = run_review_batch(config)

    print(f"JSONL: {result.raw_jsonl_path}")
    print(f"Internal Pack: {result.internal_pack_path}")
    print(f"待补充清单: {result.completion_path}")
    print(f"Excel: {result.xlsx_path}")
    print(f"QA Report: {result.qa_report_path}")
    print(f"QA Passed: {result.qa_passed}")
    if not result.qa_passed:
        failed_checks = [check["id"] for check in result.qa_report["checks"] if not check["passed"]]
        raise SystemExit(f"QA failed: {', '.join(failed_checks)}")


if __name__ == "__main__":
    main()
