import asyncio
import datetime as dt
import json
import shutil
import uuid
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user, get_run_manager, require_role
from app.api.schemas import RetentionCleanupResponse, RunCreateResponse, RunDeleteResponse, RunEventResponse, RunResponse
from app.api.serializers import event_to_response, run_to_response
from app.core.config import Settings
from app.core.paths import run_dir, uploads_dir
from app.db.base import SessionLocal, get_db
from app.db.models import Run, User
from app.platform.run_manager import RunManager


router = APIRouter(prefix="/runs", tags=["runs"])
TERMINAL_STATUSES = {"succeeded", "failed", "cancelled"}


@router.post("", response_model=RunCreateResponse, status_code=status.HTTP_201_CREATED)
async def create_run(
    line_id: str = Form(...),
    title: str = Form(""),
    config: str = Form("{}"),
    file: UploadFile = File(...),
    user: User = Depends(require_role("reviewer")),
    manager: RunManager = Depends(get_run_manager),
) -> RunCreateResponse:
    try:
        config_data = json.loads(config or "{}")
    except json.JSONDecodeError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="任务参数不是有效 JSON")

    settings = manager.settings
    _validate_upload_name(file.filename, settings)
    temp_path = await _save_upload_to_temp(file, settings)
    try:
        _validate_excel_file(temp_path)
        run_id = manager.create_run(
            line_id=line_id,
            title=title or file.filename or "未命名任务",
            config=config_data,
            created_by=user.username,
        )
        _, input_dir, _ = manager.storage.prepare_run_dirs(run_id)
        input_path = input_dir / "source.xlsx"
        shutil.copy2(temp_path, input_path)
        _update_run_input_file(run_id, input_path, db=None)
        manager.submit(run_id)
        return RunCreateResponse(run_id=run_id, status="queued")
    finally:
        temp_path.unlink(missing_ok=True)


@router.get("", response_model=list[RunResponse])
def list_runs(
    line_id: str | None = None,
    status_filter: str | None = Query(default=None, alias="status"),
    include_archived: bool = False,
    include_deleted: bool = False,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[RunResponse]:
    stmt = select(Run).order_by(Run.created_at.desc())
    if include_deleted and user.role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="只有管理员可以查看已删除任务")
    if not include_deleted:
        stmt = stmt.where(Run.deleted_at.is_(None))
    if not include_archived:
        stmt = stmt.where(Run.archived.is_(False))
    if line_id:
        stmt = stmt.where(Run.line_id == line_id)
    if status_filter:
        stmt = stmt.where(Run.status == status_filter)
    return [run_to_response(run) for run in db.scalars(stmt).all()]


@router.post("/retention/cleanup", response_model=RetentionCleanupResponse)
def cleanup_retention(
    dry_run: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin")),
    manager: RunManager = Depends(get_run_manager),
) -> RetentionCleanupResponse:
    now = dt.datetime.now(dt.timezone.utc)
    cutoff = now - dt.timedelta(days=manager.settings.run_retention_days)
    keep_last = max(manager.settings.run_retention_keep_last, 0)
    terminal_runs = list(
        db.scalars(
            select(Run)
            .where(Run.deleted_at.is_(None), Run.status.in_(TERMINAL_STATUSES))
            .order_by(Run.created_at.desc())
        )
    )
    candidates = []
    for index, run in enumerate(terminal_runs):
        if index < keep_last or run.archived:
            continue
        created_at = _as_aware_datetime(run.created_at)
        if created_at < cutoff:
            candidates.append(run)
    if not dry_run:
        for run in candidates:
            run.archived = True
            run.archived_at = now
        db.commit()
    return RetentionCleanupResponse(
        dry_run=dry_run,
        archived_count=len(candidates),
        candidate_run_ids=[run.run_id for run in candidates],
    )


@router.get("/{run_id}", response_model=RunResponse)
def get_run(run_id: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)) -> RunResponse:
    run = db.scalar(select(Run).where(Run.run_id == run_id, Run.deleted_at.is_(None)))
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Run not found")
    return run_to_response(run)


@router.post("/{run_id}/archive", response_model=RunResponse)
def archive_run(
    run_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("reviewer")),
) -> RunResponse:
    run = _get_visible_run_or_404(db, run_id)
    _ensure_terminal(run)
    run.archived = True
    run.archived_at = dt.datetime.now(dt.timezone.utc)
    db.commit()
    db.refresh(run)
    return run_to_response(run)


@router.post("/{run_id}/unarchive", response_model=RunResponse)
def unarchive_run(
    run_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("reviewer")),
) -> RunResponse:
    run = _get_visible_run_or_404(db, run_id)
    run.archived = False
    run.archived_at = None
    db.commit()
    db.refresh(run)
    return run_to_response(run)


@router.delete("/{run_id}", response_model=RunDeleteResponse)
def delete_run(
    run_id: str,
    delete_files: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("admin")),
    manager: RunManager = Depends(get_run_manager),
) -> RunDeleteResponse:
    run = _get_visible_run_or_404(db, run_id)
    _ensure_terminal(run)
    files_deleted = False
    if delete_files:
        files_deleted = _delete_run_files(manager.settings, run_id) and manager.artifact_store.delete_for_run(db, run_id)
    run.deleted_at = dt.datetime.now(dt.timezone.utc)
    run.archived = True
    if run.archived_at is None:
        run.archived_at = run.deleted_at
    db.commit()
    return RunDeleteResponse(run_id=run_id, deleted=True, files_deleted=files_deleted)


@router.post("/{run_id}/cancel", response_model=RunResponse)
def cancel_run(
    run_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(require_role("reviewer")),
    manager: RunManager = Depends(get_run_manager),
) -> RunResponse:
    manager.cancel(run_id)
    run = db.scalar(select(Run).where(Run.run_id == run_id))
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Run not found")
    db.refresh(run)
    return run_to_response(run)


@router.get("/{run_id}/events", response_model=list[RunEventResponse])
def get_events(
    run_id: str,
    after_id: int = 0,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    manager: RunManager = Depends(get_run_manager),
) -> list[RunEventResponse]:
    _ensure_run_exists(db, run_id)
    return [event_to_response(event) for event in manager.event_bus.get_events(db, run_id, after_id=after_id)]


@router.get("/{run_id}/events/stream")
async def stream_events(
    run_id: str,
    after_id: int = 0,
    last_event_id: str | None = Header(default=None, alias="Last-Event-ID"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    manager: RunManager = Depends(get_run_manager),
) -> StreamingResponse:
    _ensure_run_exists(db, run_id)
    if last_event_id and not after_id:
        try:
            after_id = int(last_event_id)
        except ValueError:
            after_id = 0

    async def event_generator():
        last_id = after_id
        while True:
            with SessionLocal() as event_db:
                events = manager.event_bus.get_events(event_db, run_id, after_id=last_id, limit=100)
                current_run = event_db.scalar(select(Run).where(Run.run_id == run_id))
                terminal = bool(current_run and current_run.status in TERMINAL_STATUSES)
            for event in events:
                last_id = event.id
                data = event_to_response(event).model_dump_json()
                yield f"id: {event.id}\nevent: {event.event_type}\ndata: {data}\n\n"
            if terminal and not events:
                break
            await asyncio.sleep(0.75)

    return StreamingResponse(event_generator(), media_type="text/event-stream")


def _ensure_run_exists(db: Session, run_id: str) -> None:
    if not db.scalar(select(Run.id).where(Run.run_id == run_id, Run.deleted_at.is_(None))):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Run not found")


def _get_visible_run_or_404(db: Session, run_id: str) -> Run:
    run = db.scalar(select(Run).where(Run.run_id == run_id, Run.deleted_at.is_(None)))
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Run not found")
    return run


def _ensure_terminal(run: Run) -> None:
    if run.status not in TERMINAL_STATUSES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="运行中任务不能归档或删除，请先等待完成或取消。")


def _delete_run_files(settings: Settings, run_id: str) -> bool:
    path = run_dir(settings, run_id)
    if not path.exists():
        return True
    try:
        shutil.rmtree(path)
        return True
    except OSError:
        return False


def _as_aware_datetime(value: dt.datetime) -> dt.datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=dt.timezone.utc)
    return value.astimezone(dt.timezone.utc)


def _validate_upload_name(filename: str | None, settings: Settings) -> None:
    suffix = Path(filename or "").suffix.lower()
    allowed = {extension.lower() for extension in settings.upload_allowed_extensions}
    if suffix not in allowed:
        allowed_text = "、".join(sorted(allowed))
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"只支持 {allowed_text} 格式，请把源数据另存为 Excel 工作簿后再上传。",
        )


async def _save_upload_to_temp(file: UploadFile, settings: Settings) -> Path:
    temp_dir = uploads_dir(settings)
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_path = temp_dir / f"upload_{uuid.uuid4().hex}.xlsx"
    total = 0
    try:
        with temp_path.open("wb") as target:
            while chunk := await file.read(1024 * 1024):
                total += len(chunk)
                if total > settings.upload_max_bytes:
                    raise HTTPException(
                        status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                        detail=f"文件超过上传上限 {settings.upload_max_bytes // 1024 // 1024} MB，请拆分后再上传。",
                    )
                target.write(chunk)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
    finally:
        await file.close()
    if total <= 0:
        temp_path.unlink(missing_ok=True)
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="文件为空，请重新选择源数据 Excel。")
    return temp_path


def _validate_excel_file(path: Path) -> None:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (InvalidFileException, OSError, ValueError):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel 文件打不开，请确认文件没有损坏，并使用 .xlsx 工作簿格式。",
        )
    try:
        visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        if not visible_sheets:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel 至少需要包含一个可见工作表。")
        first_sheet = visible_sheets[0]
        if first_sheet.max_row < 2 or first_sheet.max_column < 1:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel 至少需要包含表头和一行数据，请确认上传的是源数据表。",
            )
    finally:
        workbook.close()


def _update_run_input_file(run_id: str, input_path: Path, db: Session | None = None) -> None:
    payload = [{"name": input_path.name, "path": str(input_path), "size_bytes": input_path.stat().st_size}]
    if db is not None:
        run = db.scalar(select(Run).where(Run.run_id == run_id))
        if run:
            run.input_files_json = json.dumps(payload, ensure_ascii=False)
        return
    from app.db.base import SessionLocal

    with SessionLocal() as session:
        run = session.scalar(select(Run).where(Run.run_id == run_id))
        if run:
            run.input_files_json = json.dumps(payload, ensure_ascii=False)
            session.commit()
